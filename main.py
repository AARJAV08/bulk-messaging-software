import openpyxl
import pywhatkit as kit
import speech_recognition as sr
import time
import tkinter as tk
from tkinter import filedialog

def read_excel_data(file_path, sheet_name, phone_col):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return []

    if phone_col < 1 or phone_col > sheet.max_column:
        print(f"Error: Invalid phone column number ({phone_col}).")
        return []

    phone_numbers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if phone_col == len(row):
            phone_number = row[phone_col - 1]  # Assuming phone_col is 1-based
            phone_numbers.append(phone_number)
        else:
            print(f"Warning:could not find the contact numbers in the given column")
    wb.close()
    return phone_numbers

def speech_to_text_message():
    recognizer = sr.Recognizer()
    with sr.Microphone() as m:
        audio = recognizer.listen(m)

    try:
        message_label.config(text="Listening...")
        message = recognizer.recognize_google(audio)
        print("Your message:", message)
        text_message.delete("1.0", tk.END)  # Clear the existing text in the text area
        text_message.insert("1.0", message)  # Insert the recognized message into the text area
        message_label.config(text="Message received!")
    except sr.UnknownValueError:
        print("Sorry, could not understand audio.")
        message_label.config(text="Sorry, could not understand audio.")
    except sr.RequestError as e:
        print("Error occurred while handling the request: {0}".format(e))
        message_label.config(text="Error occurred while handling the request.")


def send_bulk_whatsapp_messages(phone_numbers, message, delay=0):
    for phone_number in phone_numbers:
        try:
            print(f"Sending message to: {phone_number}")
            kit.sendwhatmsg_instantly(phone_number, message)
            print("Message sent successfully.")
            time.sleep(delay)
        except Exception as e:
            print(f"Error sending message to {phone_number}: {e}")

#------------------------------
def browse_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)

def send_messages():
    excel_file_path = excel_file_entry.get()
    sheet_name = sheet_name_entry.get()
    phone_column = int(phone_column_entry.get())

    phone_numbers_to_send = read_excel_data(excel_file_path, sheet_name, phone_column)
    if not phone_numbers_to_send:
        result_label.config(text="Error: No phone numbers found.")
        return

    message = text_message.get("1.0", tk.END).strip()

    if not message:
        result_label.config(text="Error: Message cannot be empty.")
        return

    send_bulk_whatsapp_messages(phone_numbers_to_send, message)
    result_label.config(text="Messages sent successfully.")


# Set WhatsApp-like colors
whatsapp_green = "#25D366"
whatsapp_light_green = "#DCF8C6"
whatsapp_dark_gray = "#1E1E1E"
whatsapp_light_gray = "#ECE5DD"
whatsapp_blue = "#34B7F1"
whatsapp_dark_blue = "#0B6EFD"

#define colours
colour0 = "#1E1E1E"
colour1 = "white"

# Initial theme mode (dark mode)
current_theme = "dark"

def toggle_theme():
    global current_theme
    if current_theme == "dark":
        root.configure(bg=whatsapp_light_gray)
        excel_file_label.config(bg = whatsapp_light_gray,fg = colour0)
        sheet_name_label.config(bg = whatsapp_light_gray,fg = colour0)
        phone_column_label.config(bg = whatsapp_light_gray,fg = colour0)
        text_message_label.config(bg = whatsapp_light_gray,fg = colour0)
        message_label.config(bg=whatsapp_light_gray,fg=colour0)
        result_label.config(bg=whatsapp_light_gray,fg=colour0)
        send_button.config(bg=whatsapp_green, fg="white")
        theme_button.config(text="dark mode", bg=colour0, fg=colour1)
        current_theme = "light"
    else:
        root.configure(bg=colour0)
        excel_file_label.config(bg = colour0,fg = colour1)
        sheet_name_label.config(bg = colour0,fg = colour1)
        phone_column_label.config(bg = colour0,fg = colour1)
        text_message_label.config(bg = colour0,fg = colour1)
        message_label.config(bg=colour0,fg=colour1)
        result_label.config(bg=colour0,fg=colour1)
        send_button.config(bg=whatsapp_green, fg="white")
        theme_button.config(text="light mode", bg=colour1, fg=colour0)
        current_theme = "dark"

# Create the main window
root = tk.Tk()
root.title("Bulk WhatsApp Message Sender")

# Set the initial dark mode theme
root.configure(bg=colour0)

# Create GUI elements
excel_file_label = tk.Label(root, text="Excel File Path:", bg=colour0, fg=whatsapp_light_gray)
excel_file_entry = tk.Entry(root, width=50)
browse_button = tk.Button(root, text="Browse", command=browse_excel_file, bg=whatsapp_green, fg="white")

sheet_name_label = tk.Label(root, text="Sheet Name:", bg=colour0, fg=whatsapp_light_gray)
sheet_name_entry = tk.Entry(root, width=20)

phone_column_label = tk.Label(root, text="Phone Column:", bg=colour0, fg=whatsapp_light_gray)
phone_column_entry = tk.Entry(root, width=5)

message_label = tk.Label(root, text="", bg=colour0, fg=colour1)

send_button = tk.Button(root, text="Send Messages", command=send_messages, bg=whatsapp_green, fg="white")
result_label = tk.Label(root, text="", bg=colour0, fg = colour1)

# Create the text area and button for fetching audio in text
text_message_label = tk.Label(root, text="Text Message:", bg=colour0, fg=whatsapp_light_gray)
text_message = tk.Text(root, wrap=tk.WORD, width=50, height=10)
speech_to_text_button = tk.Button(root, text="speech-to-text", command=speech_to_text_message, bg=whatsapp_green, fg="white")

# Grid layout for GUI elements
excel_file_label.grid(row=1, column=0, padx=5, pady=5)
excel_file_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=5)
browse_button.grid(row=1, column=3, padx=5, pady=5)

sheet_name_label.grid(row=2, column=0, padx=5, pady=5)
sheet_name_entry.grid(row=2, column=1, padx=5, pady=5)

phone_column_label.grid(row=2, column=2, padx=5, pady=5)
phone_column_entry.grid(row=2, column=3, padx=5, pady=5)

text_message_label.grid(row=3, column=0, padx=5, pady=5)
text_message.grid(row=5, column=1, columnspan=2, padx=5, pady=5)
speech_to_text_button.grid(row=5, column=3, padx=5, pady=5)

send_button.grid(row=6, column=0, columnspan=4, padx=5, pady=10)
message_label.grid(row=7, column=0, columnspan=4, padx=5, pady=5)
result_label.grid(row=8, column=0, columnspan=4, padx=5, pady=5)

# Create a button to toggle the theme
theme_button = tk.Button(root, text="light mode", command=toggle_theme, bg= colour1 , fg= colour0 )
theme_button.grid(row=0, column=0, padx=5, pady=5)

# Start the main event loop
root.mainloop()
